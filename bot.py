import asyncio
import logging
import os
import sys
import threading
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from aiogram import Bot, Dispatcher, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Command
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery
from flask import Flask

# ================== НАСТРОЙКИ ==================
BOT_TOKEN = "8574535040:AAHqkLpcFgJ9xq8rPWjMehs5r1bXlbfB00s"

# Дни недели
WEEKDAYS = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ================== ПРОВЕРКА ФАЙЛА РАСПИСАНИЯ ==================
print("=" * 60)
print("🔍 ДИАГНОСТИКА:")
print(f"📁 Текущая директория: {os.getcwd()}")
print(f"📄 Список файлов в текущей директории:")

try:
    files = os.listdir('.')
    for file in files:
        if os.path.isfile(file):
            size = os.path.getsize(file)
            print(f"   - {file} ({size} bytes)")
        else:
            print(f"   - {file}/ (папка)")
except Exception as e:
    print(f"❌ Ошибка при чтении директории: {e}")

# Проверяем наличие файла расписания
schedule_found = False
schedule_path = None

if os.path.exists('schedule.xlsx'):
    schedule_found = True
    schedule_path = 'schedule.xlsx'
    print("✅ Файл schedule.xlsx НАЙДЕН в корневой папке!")
else:
    print("❌ Файл schedule.xlsx НЕ НАЙДЕН в корневой папке!")
    
    # Ищем файл во всех подпапках
    print("🔍 Поиск файла schedule.xlsx во всех папках...")
    for root, dirs, files in os.walk('.'):
        if 'schedule.xlsx' in files:
            full_path = os.path.join(root, 'schedule.xlsx')
            print(f"✅ Файл найден в: {full_path}")
            schedule_found = True
            schedule_path = full_path
            break

if not schedule_found:
    print("⚠️ Файл schedule.xlsx не найден! Создаю файл с расписанием...")
    
    try:
        wb = Workbook()
        ws = wb.active
        
        # Заголовки
        headers = ["Группа", "Дата", "День", "Время", "Аудитория", "Тип", "Предмет", "Преподаватель"]
        ws.append(headers)
        
        # Данные расписания
        data = [
            # Психология (37.03.01)
            ["Психология (37.03.01)", "02.03.2026", "Понедельник", "18:40 - 20:15", "4-541", "Лекционные занятия", "Гражданское население в противодействии распространению идеологии терроризма", "Овчинников Д.А."],
            ["Психология (37.03.01)", "03.03.2026", "Вторник", "10:25 - 12:00", "3-107", "Лекционные занятия", "Оказание первой помощи", "Корнеева Л.Н."],
            ["Психология (37.03.01)", "03.03.2026", "Вторник", "12:40 - 14:15", "Спортзал", "Элективные дисциплины", "Физическая культура и спорт", "Цветкова Л.Н., Орехов А.И., Девятиярова Е.А., Михалева О.В., Солнцева Н.С., Шестакова Т.А."],
            ["Психология (37.03.01)", "03.03.2026", "Вторник", "14:25 - 16:00", "4-519", "Лекционные занятия", "Психология развития и возрастная психология", "Шелиспанская Э.В."],
            ["Психология (37.03.01)", "04.03.2026", "Среда", "12:40 - 14:15", "4-540", "Лекционные занятия", "Деловая коммуникация в профессиональной деятельности", "Яковлева А.А."],
            ["Психология (37.03.01)", "04.03.2026", "Среда", "14:25 - 16:00", "4-540", "Практические занятия", "Учебная ознакомительная практика", "Бобровникова Н.С."],
            ["Психология (37.03.01)", "04.03.2026", "Среда", "16:10 - 17:45", "4-540", "Практические занятия", "Общая психология с практикумом", "Бобровникова Н.С."],
            ["Психология (37.03.01)", "05.03.2026", "Четверг", "08:40 - 10:15", "4-519", "Лекционные занятия", "Психолингвистика", "Чихачева Т.А."],
            ["Психология (37.03.01)", "05.03.2026", "Четверг", "10:25 - 12:00", "4-530", "Практические занятия", "Психология развития и возрастная психология", "Бобровникова Н.С."],
            ["Психология (37.03.01)", "05.03.2026", "Четверг", "12:40 - 14:15", "4-541", "Лекционные занятия", "Основы российской государственности", "Солопов О.В."],
            ["Психология (37.03.01)", "05.03.2026", "Четверг", "14:25 - 16:00", "4-519", "Практические занятия", "Основы российской государственности", "Солопов О.В."],
            ["Психология (37.03.01)", "06.03.2026", "Пятница", "08:40 - 10:15", "2-306", "Практические занятия", "Иностранный язык (немецкий)", "Бессонова Н.В."],
            ["Психология (37.03.01)", "06.03.2026", "Пятница", "08:40 - 10:15", "4-540", "Практические занятия", "Иностранный язык (английский)", "Волкова Е.В."],
            ["Психология (37.03.01)", "06.03.2026", "Пятница", "10:25 - 12:00", "2-306", "Практические занятия", "Иностранный язык (немецкий)", "Бессонова Н.В."],
            ["Психология (37.03.01)", "06.03.2026", "Пятница", "10:25 - 12:00", "4-540", "Практические занятия", "Иностранный язык (английский)", "Кудинова О.А."],
            ["Психология (37.03.01)", "07.03.2026", "Суббота", "10:25 - 12:00", "4-519", "Практические занятия", "Безопасность жизнедеятельности", "Архипцева М.С."],
            ["Психология (37.03.01)", "07.03.2026", "Суббота", "12:40 - 14:15", "4-540", "Практические занятия", "Безопасность жизнедеятельности", "Архипцева М.С."],
            ["Психология (37.03.01)", "07.03.2026", "Суббота", "12:40 - 14:15", "4-541", "Практические занятия", "Основы российской государственности", "Солопов О.В."],
            ["Психология (37.03.01)", "07.03.2026", "Суббота", "14:25 - 16:00", "4-540", "Лекционные занятия", "Социальная психология с практикумом", "Шалагинова К.С."],
            ["Психология (37.03.01)", "07.03.2026", "Суббота", "16:10 - 17:45", "4-519", "Лекционные занятия", "Психология конфликта", "Шалагинова К.С."],
            
            # Психолого-педагогическое образование (44.03.02)
            ["Психолого-педагогическое образование (44.03.02)", "02.03.2026", "Понедельник", "10:25 - 12:00", "4-528а", "Лекционные занятия", "Психология конфликта", "Шалагинова К.С."],
            ["Психолого-педагогическое образование (44.03.02)", "02.03.2026", "Понедельник", "12:40 - 14:15", "4-528а", "Лекционные занятия", "Введение в логопедическую специальность", "Лещенко С.Г."],
            ["Психолого-педагогическое образование (44.03.02)", "02.03.2026", "Понедельник", "14:25 - 16:00", "4-540", "Лекционные занятия", "Психология развития и возрастная психология", "Пазухина С.В."],
            ["Психолого-педагогическое образование (44.03.02)", "03.03.2026", "Вторник", "10:25 - 12:00", "4-541", "Лекционные занятия", "Психология общения и речевые практики", "Залыгаева С.А."],
            ["Психолого-педагогическое образование (44.03.02)", "03.03.2026", "Вторник", "12:40 - 14:15", "4-540", "Практические занятия", "Педагогика", "Катков Д.Р."],
            ["Психолого-педагогическое образование (44.03.02)", "03.03.2026", "Вторник", "14:25 - 16:00", "4-116", "Практические занятия", "Педагогика", "Яковлева А.А."],
            ["Психолого-педагогическое образование (44.03.02)", "04.03.2026", "Среда", "08:40 - 10:15", "4-540", "Лекционные занятия", "Общая психология с практикумом", "Хвалина Н.А."],
            ["Психолого-педагогическое образование (44.03.02)", "04.03.2026", "Среда", "10:25 - 12:00", "4-514", "Лекционные занятия", "Психология развития и возрастная психология", "Хвалина Н.А."],
            ["Психолого-педагогическое образование (44.03.02)", "04.03.2026", "Среда", "12:40 - 14:15", "4-514", "Лекционные занятия", "Педагогика", "Федотенко И.Л."],
            ["Психолого-педагогическое образование (44.03.02)", "04.03.2026", "Среда", "14:25 - 16:00", "4-538", "Практические занятия", "Введение в логопедическую специальность", "Шопина С.А."],
            ["Психолого-педагогическое образование (44.03.02)", "04.03.2026", "Среда", "16:10 - 17:45", "4-538", "Практические занятия", "Учебная ознакомительная практика", "Шопина С.А."],
            ["Психолого-педагогическое образование (44.03.02)", "05.03.2026", "Четверг", "10:25 - 12:00", "2-75", "Лекционные занятия", "Основы нейропсихологии", "Красникова И.В."],
            ["Психолого-педагогическое образование (44.03.02)", "05.03.2026", "Четверг", "12:40 - 14:15", "4-538", "Практические занятия", "Возрастная психология", "Требушенкова О.Ю."],
            ["Психолого-педагогическое образование (44.03.02)", "06.03.2026", "Пятница", "08:40 - 10:15", "4-540", "Практические занятия", "Иностранный язык (английский)", "Волкова Е.В."],
            ["Психолого-педагогическое образование (44.03.02)", "06.03.2026", "Пятница", "10:25 - 12:00", "4-521", "Практические занятия", "Безопасность жизнедеятельности", "Архипцева М.С."],
            ["Психолого-педагогическое образование (44.03.02)", "06.03.2026", "Пятница", "12:40 - 14:15", "4-540", "Практические занятия", "Безопасность жизнедеятельности", "Архипцева М.С."],
            
            # Специальное (дефектологическое) образование (44.03.03)
            ["Специальное (дефектологическое) образование (44.03.03)", "02.03.2026", "Понедельник", "10:25 - 12:00", "4-519", "Лекционные занятия", "Социальная психология", "Шелиспанская Э.В."],
            ["Специальное (дефектологическое) образование (44.03.03)", "02.03.2026", "Понедельник", "14:25 - 16:00", "4-519", "Лекционные занятия", "Социальная педагогика", "Карандеева А.В."],
            ["Специальное (дефектологическое) образование (44.03.03)", "03.03.2026", "Вторник", "08:40 - 10:15", "4-541", "Лекционные занятия", "Безопасность жизнедеятельности", "Петрова М.С."],
            ["Специальное (дефектологическое) образование (44.03.03)", "03.03.2026", "Вторник", "12:40 - 14:15", "4-541", "Практические занятия", "Безопасность жизнедеятельности", "Архипцева М.С."],
            ["Специальное (дефектологическое) образование (44.03.03)", "03.03.2026", "Вторник", "14:25 - 16:00", "4-530", "Практические занятия", "Безопасность жизнедеятельности", "Архипцева М.С."],
            ["Специальное (дефектологическое) образование (44.03.03)", "04.03.2026", "Среда", "08:40 - 10:15", "2-69", "Лекционные занятия", "Невропатология", "Маркова М.П."],
            ["Специальное (дефектологическое) образование (44.03.03)", "04.03.2026", "Среда", "10:25 - 12:00", "2-69", "Лекционные занятия", "Невропатология", "Маркова М.П."],
            ["Специальное (дефектологическое) образование (44.03.03)", "04.03.2026", "Среда", "12:40 - 14:15", "4-538", "Практические занятия", "Психология общения и речевые практики", "Бобровникова Н.С."],
            ["Специальное (дефектологическое) образование (44.03.03)", "04.03.2026", "Среда", "16:10 - 17:45", "4-519", "Лекционные занятия", "Социальная психология", "Шелиспанская Э.В."],
            ["Специальное (дефектологическое) образование (44.03.03)", "05.03.2026", "Четверг", "10:25 - 12:00", "4-519", "Практические занятия", "Иностранный язык (английский)", "Данилова Ю.С."],
            ["Специальное (дефектологическое) образование (44.03.03)", "05.03.2026", "Четверг", "14:25 - 16:00", "4-530", "Практические занятия", "Социальная психология", "Шалагинова К.С."],
            ["Специальное (дефектологическое) образование (44.03.03)", "06.03.2026", "Пятница", "08:40 - 10:15", "4-519", "Практические занятия", "Психолингвистика", "Корабельникова Е.И."],
            ["Специальное (дефектологическое) образование (44.03.03)", "06.03.2026", "Пятница", "12:40 - 14:15", "4-519", "Практические занятия", "Социальная педагогика", "Карандеева А.В."],
            
            # Педагогика и психология девиантного поведения (44.05.01)
            ["Педагогика и психология девиантного поведения (44.05.01)", "02.03.2026", "Понедельник", "10:25 - 12:00", "4-519", "Лекционные занятия", "Психология конфликта", "Шалагинова К.С."],
            ["Педагогика и психология девиантного поведения (44.05.01)", "02.03.2026", "Понедельник", "14:25 - 16:00", "4-519", "Лекционные занятия", "Социальная педагогика", "Карандеева А.В."],
            ["Педагогика и психология девиантного поведения (44.05.01)", "03.03.2026", "Вторник", "10:25 - 12:00", "4-519", "Лекционные занятия", "Психология личности педагога и обучающегося", "Куликова Т.И."],
            ["Педагогика и психология девиантного поведения (44.05.01)", "03.03.2026", "Вторник", "12:40 - 14:15", "4-116", "Лекционные занятия", "Возрастная психология", "Кацеро А.А."],
            ["Педагогика и психология девиантного поведения (44.05.01)", "03.03.2026", "Вторник", "14:25 - 16:00", "2-77", "Лекционные занятия", "Нейрофизиология", "Красников Г.В."],
            ["Педагогика и психология девиантного поведения (44.05.01)", "03.03.2026", "Вторник", "16:10 - 17:45", "2-77", "Лекционные занятия", "Нейрофизиология", "Красников Г.В."],
            ["Педагогика и психология девиантного поведения (44.05.01)", "04.03.2026", "Среда", "08:40 - 10:15", "4-514", "Лекционные занятия", "Социальная психология", "Шалагинова К.С."],
            ["Педагогика и психология девиантного поведения (44.05.01)", "04.03.2026", "Среда", "10:25 - 12:00", "4-514", "Лекционные занятия", "Социальная психология", "Шалагинова К.С."],
            ["Педагогика и психология девиантного поведения (44.05.01)", "04.03.2026", "Среда", "12:40 - 14:15", "4-519", "Лекционные занятия", "Психологическое консультирование по проблемам девиантного поведения", "Шелиспанская Э.В."],
            ["Педагогика и психология девиантного поведения (44.05.01)", "04.03.2026", "Среда", "14:25 - 16:00", "4-519", "Лекционные занятия", "Психологическое консультирование по проблемам девиантного поведения", "Шелиспанская Э.В."],
            ["Педагогика и психология девиантного поведения (44.05.01)", "05.03.2026", "Четверг", "10:25 - 12:00", "4-519", "Лекционные занятия", "Основы нейропсихологии", "Красникова И.В."],
            ["Педагогика и психология девиантного поведения (44.05.01)", "05.03.2026", "Четверг", "14:25 - 16:00", "2-75", "Лекционные занятия", "Основы нейропсихологии", "Красникова И.В."],
            ["Педагогика и психология девиантного поведения (44.05.01)", "06.03.2026", "Пятница", "10:25 - 12:00", "4-519", "Практические занятия", "Психология общения и речевые практики", "Подкопаева Е.С."],
            ["Педагогика и психология девиантного поведения (44.05.01)", "06.03.2026", "Пятница", "14:25 - 16:00", "4-519", "Практические занятия", "Основы нейропсихологии", "Красникова И.В."],
            ["Педагогика и психология девиантного поведения (44.05.01)", "07.03.2026", "Суббота", "08:40 - 10:15", "4-519", "Практические занятия", "Учебная ознакомительная практика", "Бобровикова Н.С."],
            ["Педагогика и психология девиантного поведения (44.05.01)", "07.03.2026", "Суббота", "12:40 - 14:15", "2-108", "Лабораторные занятия", "Оказание первой помощи", "Ибрагимова Т.А."],
            ["Педагогика и психология девиантного поведения (44.05.01)", "07.03.2026", "Суббота", "14:25 - 16:00", "2-108", "Лабораторные занятия", "Оказание первой помощи", "Ибрагимова Т.А."],
        ]
        
        # Записываем данные
        for row in data:
            ws.append(row)
        
        # Автоматическая ширина колонок
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Сохраняем файл
        wb.save('schedule.xlsx')
        print(f"✅ ФАЙЛ РАСПИСАНИЯ СОЗДАН! Записей: {len(data)}")
        schedule_found = True
        schedule_path = 'schedule.xlsx'
    except Exception as e:
        print(f"❌ ОШИБКА ПРИ СОЗДАНИИ ФАЙЛА: {e}")

print("=" * 60)

# ================== РАБОТА С EXCEL ==================
class ScheduleDatabase:
    def __init__(self):
        self.schedule = []
        self.groups = []
        self.load_data()
    
    def load_data(self):
        """Загрузить данные из Excel файла"""
        try:
            # Используем найденный путь к файлу
            file_path = schedule_path if schedule_path else 'schedule.xlsx'
            logger.info(f"📂 Загрузка файла: {file_path}")
            
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Читаем заголовки
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            logger.info(f"📋 Заголовки: {headers}")
            
            # Читаем данные
            self.schedule = []
            groups_set = set()
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # если есть группа
                    record = {}
                    for i, header in enumerate(headers):
                        record[header] = row[i] if i < len(row) else ''
                    self.schedule.append(record)
                    groups_set.add(row[0])
            
            self.groups = sorted(list(groups_set))
            logger.info(f"✅ Загружено {len(self.schedule)} записей из Excel")
            logger.info(f"📋 Найдены группы: {', '.join(self.groups)}")
            return True
        except Exception as e:
            logger.error(f"❌ Ошибка загрузки Excel: {e}")
            return False
    
    def get_groups_list(self) -> list:
        return self.groups
    
    def get_schedule_for_date(self, group: str, target_date: datetime) -> list:
        """Получить расписание на конкретную дату"""
        result = []
        target_date_str = target_date.strftime('%d.%m.%Y')
        
        for lesson in self.schedule:
            if lesson.get('Группа') == group and lesson.get('Дата') == target_date_str:
                result.append(lesson)
        
        return sorted(result, key=lambda x: x.get('Время', ''))
    
    def get_schedule_for_week(self, group: str, start_date: datetime) -> dict:
        """Получить расписание на неделю"""
        # Определяем границы недели
        monday = start_date - timedelta(days=start_date.weekday())
        sunday = monday + timedelta(days=6)
        
        weekly = {day: [] for day in WEEKDAYS}
        
        for lesson in self.schedule:
            if lesson.get('Группа') != group:
                continue
            
            # Парсим дату занятия
            try:
                lesson_date = datetime.strptime(lesson.get('Дата', ''), '%d.%m.%Y')
                if monday <= lesson_date <= sunday:
                    day_name = lesson.get('День')
                    if day_name in weekly:
                        weekly[day_name].append(lesson)
            except:
                continue
        
        # Сортируем по времени
        for day in weekly:
            weekly[day] = sorted(weekly[day], key=lambda x: x.get('Время', ''))
        
        return weekly

# ================== СОСТОЯНИЯ ==================
class ScheduleStates(StatesGroup):
    waiting_for_group = State()
    viewing_schedule = State()

# ================== КЛАВИАТУРЫ ==================
def get_main_keyboard() -> ReplyKeyboardMarkup:
    """Главная клавиатура"""
    buttons = [
        [KeyboardButton(text="📅 Сегодня"), KeyboardButton(text="📆 Завтра")],
        [KeyboardButton(text="📋 Эта неделя"), KeyboardButton(text="📌 Следующая неделя")],
        [KeyboardButton(text="🔍 Выбрать группу")]
    ]
    return ReplyKeyboardMarkup(resize_keyboard=True, keyboard=buttons)

def get_groups_keyboard(groups: list) -> InlineKeyboardMarkup:
    """Клавиатура со списком групп"""
    if not groups:
        return None
    
    buttons = []
    for i in range(0, len(groups), 2):
        row = []
        row.append(InlineKeyboardButton(text=groups[i], callback_data=f"group_{i}"))
        if i + 1 < len(groups):
            row.append(InlineKeyboardButton(text=groups[i + 1], callback_data=f"group_{i+1}"))
        buttons.append(row)
    
    return InlineKeyboardMarkup(inline_keyboard=buttons)

# ================== ИНИЦИАЛИЗАЦИЯ ==================
storage = MemoryStorage()
bot = Bot(token=BOT_TOKEN, parse_mode="HTML")
dp = Dispatcher(bot, storage=storage)
db = ScheduleDatabase()

# ================== ОБРАБОТЧИКИ ==================
@dp.message_handler(Command("start"), state="*")
async def cmd_start(message: types.Message, state: FSMContext):
    await state.finish()
    
    if not db.groups:
        await message.answer("❌ Ошибка загрузки расписания. Проверьте файл schedule.xlsx")
        return
    
    await message.answer(
        "👋 Привет! Я бот для просмотра расписания ТГПУ\n"
        "Чтобы начать, выберите вашу группу:",
        reply_markup=get_main_keyboard()
    )
    
    await message.answer(
        "📚 Доступные группы:\nВыберите вашу группу:",
        reply_markup=get_groups_keyboard(db.groups)
    )
    await ScheduleStates.waiting_for_group.set()

@dp.callback_query_handler(lambda c: c.data and c.data.startswith("group_"), state="*")
async def process_group_selection(callback: CallbackQuery, state: FSMContext):
    group_index = int(callback.data.replace("group_", ""))
    
    if 0 <= group_index < len(db.groups):
        group = db.groups[group_index]
        await state.update_data(selected_group=group)
        await ScheduleStates.viewing_schedule.set()
        
        await callback.message.edit_text(f"✅ Выбрана группа: {group}")
        await callback.message.answer(
            f"Что хотите узнать?",
            reply_markup=get_main_keyboard()
        )
    else:
        await callback.message.edit_text("❌ Ошибка выбора группы")
    
    await callback.answer()

@dp.message_handler(lambda message: message.text == "📅 Сегодня", state=ScheduleStates.viewing_schedule)
async def cmd_today(message: types.Message, state: FSMContext):
    data = await state.get_data()
    group = data.get("selected_group")
    
    if not group:
        await ScheduleStates.waiting_for_group.set()
        await message.answer("Сначала выберите группу!")
        return
    
    today = datetime.now()
    lessons = db.get_schedule_for_date(group, today)
    
    if not lessons:
        await message.answer(f"📅 {today.strftime('%d.%m.%Y')}\n___________________________________\n\nПар нет 🎉")
        return
    
    text = f"📅 {today.strftime('%d.%m.%Y')}\n___________________________________\n"
    for i, lesson in enumerate(lessons):
        text += f"\n{lesson.get('Время', '')}  📚"
        text += f"\n🏛 {lesson.get('Аудитория', '')}"
        text += f"\n📖 {lesson.get('Тип', '')}"
        text += f"\n📚 {lesson.get('Предмет', '')}"
        text += f"\n👨‍🏫 {lesson.get('Преподаватель', '')}"
        if i < len(lessons) - 1:
            text += "\n" + "-" * 40 + "\n"
    
    await message.answer(text)

@dp.message_handler(lambda message: message.text == "📆 Завтра", state=ScheduleStates.viewing_schedule)
async def cmd_tomorrow(message: types.Message, state: FSMContext):
    data = await state.get_data()
    group = data.get("selected_group")
    
    if not group:
        await ScheduleStates.waiting_for_group.set()
        await message.answer("Сначала выберите группу!")
        return
    
    tomorrow = datetime.now() + timedelta(days=1)
    lessons = db.get_schedule_for_date(group, tomorrow)
    
    if not lessons:
        await message.answer(f"📅 {tomorrow.strftime('%d.%m.%Y')}\n___________________________________\n\nПар нет 🎉")
        return
    
    text = f"📅 {tomorrow.strftime('%d.%m.%Y')}\n___________________________________\n"
    for i, lesson in enumerate(lessons):
        text += f"\n{lesson.get('Время', '')}  📚"
        text += f"\n🏛 {lesson.get('Аудитория', '')}"
        text += f"\n📖 {lesson.get('Тип', '')}"
        text += f"\n📚 {lesson.get('Предмет', '')}"
        text += f"\n👨‍🏫 {lesson.get('Преподаватель', '')}"
        if i < len(lessons) - 1:
            text += "\n" + "-" * 40 + "\n"
    
    await message.answer(text)

@dp.message_handler(lambda message: message.text == "📋 Эта неделя", state=ScheduleStates.viewing_schedule)
async def cmd_this_week(message: types.Message, state: FSMContext):
    data = await state.get_data()
    group = data.get("selected_group")
    
    if not group:
        await ScheduleStates.waiting_for_group.set()
        await message.answer("Сначала выберите группу!")
        return
    
    today = datetime.now()
    monday = today - timedelta(days=today.weekday())
    weekly = db.get_schedule_for_week(group, today)
    
    text = f"📋 Текущая неделя (с {monday.strftime('%d.%m')})\n" + "=" * 40 + "\n"
    
    has_lessons = False
    for i, day in enumerate(WEEKDAYS):
        day_lessons = weekly[day]
        current_date = monday + timedelta(days=i)
        
        if day_lessons:
            has_lessons = True
            text += f"\n📅 {day}, {current_date.strftime('%d.%m')}\n" + "-" * 30 + "\n"
            
            for lesson in day_lessons:
                text += f"\n{lesson.get('Время', '')}  📚"
                text += f"\n🏛 {lesson.get('Аудитория', '')}"
                text += f"\n📖 {lesson.get('Тип', '')}"
                text += f"\n📚 {lesson.get('Предмет', '')}"
                text += f"\n👨‍🏫 {lesson.get('Преподаватель', '')}\n"
    
    if not has_lessons:
        text += "\nНа этой неделе пар нет 🎉"
    
    await message.answer(text)

@dp.message_handler(lambda message: message.text == "📌 Следующая неделя", state=ScheduleStates.viewing_schedule)
async def cmd_next_week(message: types.Message, state: FSMContext):
    data = await state.get_data()
    group = data.get("selected_group")
    
    if not group:
        await ScheduleStates.waiting_for_group.set()
        await message.answer("Сначала выберите группу!")
        return
    
    next_week = datetime.now() + timedelta(weeks=1)
    monday = next_week - timedelta(days=next_week.weekday())
    weekly = db.get_schedule_for_week(group, next_week)
    
    text = f"📌 Следующая неделя (с {monday.strftime('%d.%m')})\n" + "=" * 40 + "\n"
    
    has_lessons = False
    for i, day in enumerate(WEEKDAYS):
        day_lessons = weekly[day]
        current_date = monday + timedelta(days=i)
        
        if day_lessons:
            has_lessons = True
            text += f"\n📅 {day}, {current_date.strftime('%d.%m')}\n" + "-" * 30 + "\n"
            
            for lesson in day_lessons:
                text += f"\n{lesson.get('Время', '')}  📚"
                text += f"\n🏛 {lesson.get('Аудитория', '')}"
                text += f"\n📖 {lesson.get('Тип', '')}"
                text += f"\n📚 {lesson.get('Предмет', '')}"
                text += f"\n👨‍🏫 {lesson.get('Преподаватель', '')}\n"
    
    if not has_lessons:
        text += "\nНа следующей неделе пар нет 🎉"
    
    await message.answer(text)

@dp.message_handler(lambda message: message.text == "🔍 Выбрать группу", state="*")
async def cmd_change_group(message: types.Message, state: FSMContext):
    await ScheduleStates.waiting_for_group.set()
    await message.answer(
        "📚 Доступные группы:\nВыберите вашу группу:",
        reply_markup=get_groups_keyboard(db.groups)
    )

@dp.message_handler(state="*")
async def handle_unknown(message: types.Message, state: FSMContext):
    current_state = await state.get_state()
    
    if current_state == ScheduleStates.viewing_schedule.state:
        await message.answer(
            "Используйте кнопки для навигации 👆",
            reply_markup=get_main_keyboard()
        )
    elif current_state == ScheduleStates.waiting_for_group.state:
        await message.answer(
            "⚠️ Сначала выберите группу через кнопку '🔍 Выбрать группу'"
        )
    else:
        await message.answer(
            "👋 Отправьте /start для начала работы"
        )

async def main():
    """Главная функция запуска бота"""
    logger.info("🚀 Запуск бота...")
    await dp.start_polling()

# ================== FLASK СЕРВЕР ДЛЯ RAILWAY ==================
app = Flask(__name__)

@app.route('/')
def home():
    return "🤖 Бот расписания ТГПУ работает!"

@app.route('/health')
def health():
    return "OK", 200

def run_bot():
    """Запуск бота в отдельном потоке"""
    asyncio.run(main())

if __name__ == "__main__":
    # Запускаем бота в фоновом потоке
    bot_thread = threading.Thread(target=run_bot, daemon=True)
    bot_thread.start()
    logger.info("✅ Бот запущен в фоновом потоке")
    
    # Запускаем Flask сервер для Railway
    port = int(os.environ.get("PORT", 8080))
    logger.info(f"🌐 Запуск Flask сервера на порту {port}")
    app.run(host="0.0.0.0", port=port)
